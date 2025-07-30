# app/utils.py
import pandas as pd
import numpy as np
import jdatetime
from datetime import datetime
from sqlalchemy import func
import re
from openpyxl import load_workbook
import os
import logging
from app.models import Settings

# دیکشنری مپینگ واحدهای اندازه‌گیری به کدهای عددی
UNIT_OF_MEASUREMENT_MAPPING = {
    'کیلوگرم': 1,
    'تن': 2,
    'حلقه': 3,
    'عدد': 4,
    'متر': 5,
    'دستگاه': 6,
    'یارد': 7,
    'کارتن': 8,
    'کیسه': 9,
    'گالن': 10,
    'گرم': 11,
    'خط تولید': 12,
    'اونس': 13,
    'فروند': 14,
    'راس': 15,
    'دز': 16,
    'متر مربع': 17,
    'لیتر': 18,
    'دست': 19,
    'جفت': 20,
    'ورق': 21,
    'دوجین': 22,
    'فوت مربع': 23,
    'قطعه': 24,
    'بیو': 25,
    'ویال': 26,
    'نخ': 27,
    'iu': 28,
    'میلی گرم': 29,
    'یونیت': 30,
    'KA': 31,
    'قراص': 32,
    'قیراط': 33,
    'میلی متر': 34,
    'نفر': 35,
    'قلاده': 36,
    'کیلو وات ساعت': 37,
    '1000 واحد': 38,
    'باکس': 39,
    'پاکت': 40,
    'مترمکعب': 41,
    'جلد': 42,
    'صفحه': 43,
    'توپ': 44,
    'ست': 45,
    'بسته': 46,
    'تخته': 47,
    'رول': 48,
    'طاقه': 49,
    'پالت': 50,
    'ثوب': 51,
    'نیم دوجین': 52,
    'قرقره': 53,
    'بطری': 54,
    'برگ': 55,
    'سطل': 56,
    'شاخه': 57,
    'قوطی': 58,
    'جلد': 59,
    'تیوب': 60,
    'کلاف': 61,
    'کیسه': 62,
    'طغرا': 63,
    'بشکه': 64,
    'کارتن (دخانیات)': 65,
    'قراصه': 66,
    'لنگه': 67,
    'عدل': 68,
    'جعبه': 69,
    'تعداد': 70,
    'سانتی متر': 71,
    'پد': 72,
    'واحد': 1000
}

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def extract_number(text):
    if pd.isna(text) or text is None:
        return ""
    return re.sub(r'\D', '', str(text))

def split_name(full_name):
    if pd.isna(full_name) or full_name is None:
        return "", ""
    name_text = re.sub(r'.*:', '', str(full_name)).strip()
    parts = name_text.split(' ', 1)
    first_name = parts[0].strip() if parts else ""
    last_name = parts[1].strip() if len(parts) > 1 else ""
    return first_name, last_name

def process_items_excel(filepath):
    """
    Reads an Excel file, processes its rows, and returns a list of item data dictionaries
    and any validation messages. This function DOES NOT interact with the database.
    """
    messages = []
    items_to_process = []
    column_mapping = {
        'شماره سند': 'document_number',
        'شماره صورتحساب': 'invoice_number_ref',
        'تاریخ سند': 'document_date',
        'فروشنده': 'seller',
        'استان فروشنده': 'seller_province',
        'نوع فعالیت': 'activity_type',
        'مبدا': 'origin',
        'طبقه کالا': 'item_category',
        'شرح کالا': 'product_description',
        'واحداندازه‌گیری': 'unit_of_measurement',
        'تعداد / مقدار کالا': 'quantity',
        'مبلغ واحد': 'unit_price',
        'مبلغ نهایی': 'final_amount',
        'شناسه کالا': 'product_id',
        'توضیحات': 'remarks',
    }
    try:
        df = pd.read_excel(filepath, dtype=str)
        df = df.rename(columns=lambda x: x.strip())
        df = df.where(pd.notna(df), None)
        logger.debug(f"Excel columns: {list(df.columns)}")

        required_excel_columns = set(column_mapping.keys())
        actual_excel_columns = set(df.columns)
        if not required_excel_columns.issubset(actual_excel_columns):
            missing = required_excel_columns - actual_excel_columns
            messages.append(
                ('danger', f"فایل اکسل ناقص است. ستون‌های زیر یافت نشدند: {', '.join(missing)}"))
            return [], messages

        for index, row in df.iterrows():
            try:
                product_id = row.get('شناسه کالا')
                if not product_id:
                    continue
                item_data = {}
                for excel_col, model_field in column_mapping.items():
                    item_data[model_field] = row.get(excel_col)
                quantity_str = item_data.get('quantity', '0')
                item_data['quantity'] = int(float(quantity_str or 0))
                unit_price_str = item_data.get('unit_price', '0.0')
                item_data['unit_price'] = float(unit_price_str or 0.0)
                jalali_date_str = item_data.get('document_date')
                gregorian_date = None
                if jalali_date_str:
                    try:
                        gregorian_date = jdatetime.datetime.strptime(
                            str(jalali_date_str), '%Y/%m/%d').togregorian().date()
                    except (ValueError, TypeError):
                        messages.append(
                            ('warning', f"ردیف {index + 2}: فرمت تاریخ سند '{jalali_date_str}' نامعتبر است و از آن صرف‌نظر شد."))
                        continue
                if gregorian_date is None:
                    messages.append(
                        ('warning', f"ردیف {index + 2}: تاریخ سند خالی است و از آن صرف‌نظر شد."))
                    continue
                item_data['document_date'] = gregorian_date
                item_data['final_amount'] = item_data['quantity'] * item_data['unit_price']
                items_to_process.append(item_data)
            except Exception as e:
                messages.append(
                    ('danger', f"خطا در پردازش ردیف {index + 2} اکسل: {e}"))
                continue
    except Exception as e:
        messages.append(('danger', f"خطا در خواندن یا پردازش فایل اکسل: {e}"))
    return items_to_process, messages

MULTIPLIER_FACTOR = 125000

def process_excel_invoices(filepath, db, Item, ItemUsageLog, current_invoice_number_start):
    """
    Processes an invoice Excel file and assigns exactly one item from Item table
    with sufficient remaining_quantity to each product, prioritizing highest unit_price.
    Commits changes to remaining_quantity after each product to ensure up-to-date inventory.
    Prefers unique items but falls back to previously used items if no new item is available.
    Uses product description from the invoice in the output.
    Updates inventory values after processing.
    """
    db.session.expire_all()
    
    output_data = []
    log_entries = []
    messages = []
    next_invoice_number = current_invoice_number_start
    required_products = []
    used_item_ids = set()

    try:
        df = pd.read_excel(filepath, header=None, dtype=str).where(pd.notna, None)
        logger.debug(f"Excel file {filepath} loaded with shape: {df.shape}")

        CELL_POSITIONS = {
            "date": (2, 26),
            "zip_code": (11, 4),
            "national_id": (9, 16),
            "buyer_name": (9, 0)
        }
        PRODUCT_COLUMNS = {
            "quantity": 4,
            "unit_price": 6,
            "discount": 12,
            "product_description": 2
        }
        PRODUCT_START_ROW_INDEX = 15

        if df.shape[0] < PRODUCT_START_ROW_INDEX or df.shape[1] < max(col for col in CELL_POSITIONS.values())[1] + 1:
            messages.append(('danger', f"فایل {os.path.basename(filepath)} خیلی کوچک است یا ساختار نادرستی دارد."))
            return pd.DataFrame(), [], next_invoice_number, messages

        def get_cell_value(position):
            row, col = position
            try:
                return df.iloc[row, col] if row < df.shape[0] and col < df.shape[1] else ""
            except Exception as e:
                logger.warning(f"Error accessing cell at {position}: {e}")
                return ""

        date_str = get_cell_value(CELL_POSITIONS["date"]) or ""
        zip_code_raw = get_cell_value(CELL_POSITIONS["zip_code"]) or ""
        national_id_raw = get_cell_value(CELL_POSITIONS["national_id"]) or ""
        buyer_name_full = get_cell_value(CELL_POSITIONS["buyer_name"]) or ""
        
        zip_code = extract_number(zip_code_raw)
        national_id = extract_number(national_id_raw)
        buyer_name, buyer_surname = split_name(buyer_name_full)

        if not date_str:
            messages.append(('warning', f"تاریخ در فایل {os.path.basename(filepath)} خالی است. از تاریخ فعلی استفاده می‌شود."))
            date_str = datetime.utcnow().date().strftime('%Y/%m/%d')
        if not zip_code:
            messages.append(('warning', f"کد پستی در فایل {os.path.basename(filepath)} خالی است."))
        if not national_id:
            messages.append(('warning', f"کد ملی در فایل {os.path.basename(filepath)} خالی است."))
        if not buyer_name_full:
            messages.append(('warning', f"نام خریدار در فایل {os.path.basename(filepath)} خالی است."))

        for row_idx in range(PRODUCT_START_ROW_INDEX, df.shape[0]):
            try:
                raw_unit_price = df.iloc[row_idx, PRODUCT_COLUMNS["unit_price"]] if PRODUCT_COLUMNS["unit_price"] < df.shape[1] else None
                unit_price_val = pd.to_numeric(raw_unit_price, errors='coerce')
                if pd.isna(unit_price_val) or unit_price_val == 0:
                    break

                raw_quantity = df.iloc[row_idx, PRODUCT_COLUMNS["quantity"]] if PRODUCT_COLUMNS["quantity"] < df.shape[1] else None
                quantity_val = pd.to_numeric(raw_quantity, errors='coerce')
                quantity_needed = int(quantity_val) if pd.notna(quantity_val) and quantity_val > 0 else 0
                if quantity_needed <= 0:
                    messages.append(('warning', f"مقدار نامعتبر یا صفر برای محصول در ردیف {row_idx + 2}"))
                    continue

                raw_discount = df.iloc[row_idx, PRODUCT_COLUMNS["discount"]] if PRODUCT_COLUMNS["discount"] < df.shape[1] else None
                discount_val = pd.to_numeric(raw_discount, errors='coerce')
                discount = float(discount_val) if pd.notna(discount_val) else 0.0

                product_description_from_invoice = str(df.iloc[row_idx, PRODUCT_COLUMNS["product_description"]] or '').strip() if PRODUCT_COLUMNS["product_description"] < df.shape[1] else ''
                if not product_description_from_invoice:
                    messages.append(('warning', f"توضیحات محصول در ردیف {row_idx + 2} خالی است."))
                    continue

                required_products.append((product_description_from_invoice, quantity_needed, unit_price_val, discount))
                logger.debug(f"Row {row_idx + 2}: Product Description: {product_description_from_invoice}, Quantity Needed: {quantity_needed}, Unit Price: {unit_price_val}, Discount: {discount}")
            except Exception as e:
                logger.warning(f"Error processing row {row_idx + 2} in {filepath}: {e}")
                continue

        if not required_products:
            messages.append(('danger', f"هیچ محصول معتبری در فایل {os.path.basename(filepath)} یافت نشد."))
            return pd.DataFrame(), [], next_invoice_number, messages

        available_items = db.session.query(Item).filter(Item.remaining_quantity > 0).order_by(Item.unit_price.desc()).all()
        logger.debug(f"تعداد آیتم‌های با موجودی مثبت: {len(available_items)}")
        for item in available_items:
            logger.debug(f"Item: {item.product_id}, Remaining Quantity: {item.remaining_quantity}, Unit Price: {item.unit_price}, Description: {item.product_description}")

        at_least_one_product_processed = False

        for product_description_from_invoice, quantity_needed, unit_price_val, discount in required_products:
            db.session.expire_all()
            logger.debug(f"Processing product '{product_description_from_invoice}' with quantity_needed={quantity_needed}")

            item_in_db = None
            for item in db.session.query(Item).filter(
                Item.remaining_quantity >= quantity_needed,
                ~Item.id.in_(used_item_ids)
            ).order_by(Item.unit_price.desc()).all():
                item_in_db = item
                break

            if not item_in_db:
                logger.debug(f"هیچ آیتم جدیدی با موجودی کافی برای '{product_description_from_invoice}' (نیاز: {quantity_needed}) یافت نشد. بررسی آیتم‌های استفاده‌شده...")
                for item in db.session.query(Item).filter(
                    Item.remaining_quantity >= quantity_needed
                ).order_by(Item.unit_price.desc()).all():
                    item_in_db = item
                    break

            if not item_in_db:
                logger.error(f"هیچ آیتمی (جدید یا استفاده‌شده) با موجودی کافی برای '{product_description_from_invoice}' (نیاز: {quantity_needed}) یافت نشد.")
                logger.debug("همه آیتم‌های موجود در دیتابیس:")
                for item in db.session.query(Item).all():
                    logger.debug(f"Item: {item.product_id}, Remaining Quantity: {item.remaining_quantity}, Unit Price: {item.unit_price}, Used: {item.id in used_item_ids}")
                messages.append(('warning', f"برای '{product_description_from_invoice}' در فایل {os.path.basename(filepath)}، هیچ کالای با موجودی کافی (نیاز: {quantity_needed}) یافت نشد. مقدار صفر تخصیص داده شد."))
                output_data.append({
                    'A': date_str, 'B': next_invoice_number, 'C': zip_code, 'D': national_id,
                    'E': buyer_name, 'F': buyer_surname, 'G': '', 'H': '', 'I': '', 'J': '',
                    'K': '', 'L': product_description_from_invoice,
                    'M': '4', 'N': 0, 'O': 'IRR', 'P': 1, 'Q': unit_price_val, 'R': discount, 'S': 0,
                })
                continue

            logger.debug(f"برای محصول '{product_description_from_invoice}'، آیتم انتخاب‌شده: {item_in_db.product_id} با موجودی {item_in_db.remaining_quantity} و قیمت واحد {item_in_db.unit_price} (استفاده‌شده: {item_in_db.id in used_item_ids})")

            total_quantity_used = quantity_needed
            item_in_db.remaining_quantity -= total_quantity_used
            used_item_ids.add(item_in_db.id)

            item_unit_price = item_in_db.unit_price if item_in_db.unit_price is not None else unit_price_val
            calculated_vat = (item_unit_price * total_quantity_used) / 10.0
            unit_of_measurement = item_in_db.unit_of_measurement or 'عدد'
            unit_code = UNIT_OF_MEASUREMENT_MAPPING.get(unit_of_measurement.strip(), 4)

            output_data.append({
                'A': date_str, 'B': next_invoice_number, 'C': zip_code, 'D': national_id,
                'E': buyer_name, 'F': buyer_surname, 'G': '', 'H': '', 'I': '', 'J': '',
                'K': item_in_db.product_id, 'L': product_description_from_invoice,
                'M': unit_code, 'N': total_quantity_used,
                'O': 'IRR', 'P': 1, 'Q': item_unit_price, 'R': discount, 'S': calculated_vat,
            })

            pd_exit_date = pd.to_datetime(date_str, errors='coerce')
            exit_date_obj = pd_exit_date.date() if pd.notna(pd_exit_date) else datetime.utcnow().date()
            log_entries.append(ItemUsageLog(
                item_id=item_in_db.id,
                exit_date=exit_date_obj,
                invoice_number_used=str(next_invoice_number),
                quantity_used=total_quantity_used,
                price_at_usage=item_unit_price
            ))

            try:
                db.session.add(item_in_db)
                db.session.commit()
                logger.debug(f"Committed changes for item {item_in_db.product_id}: new remaining_quantity={item_in_db.remaining_quantity}")
                # محاسبه و به‌روزرسانی مقادیر ارز پس از هر تخصیص
                initial_value, remaining_value, used_value = calculate_inventory_values(db, Item, Settings)
                logger.debug(f"Updated inventory values after processing '{product_description_from_invoice}': Initial={initial_value}, Remaining={remaining_value}, Used={used_value}")
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error committing changes for item {item_in_db.product_id}: {str(e)}")
                messages.append(('danger', f"خطا در به‌روزرسانی موجودی برای '{product_description_from_invoice}': {str(e)}"))
                continue

            at_least_one_product_processed = True

        if at_least_one_product_processed:
            next_invoice_number += 1
            db.session.commit()
            initial_value, remaining_value, used_value = calculate_inventory_values(db, Item, Settings)
            messages.append(('success', f"فایل {os.path.basename(filepath)} با موفقیت پردازش شد. ارز اولیه: {initial_value:,.2f}, ارز باقیمانده: {remaining_value:,.2f}, ارز مصرف‌شده: {used_value:,.2f}"))
        else:
            messages.append(('danger', f"هیچ محصولی در فایل {os.path.basename(filepath)} قابل پردازش نبود."))
            return pd.DataFrame(), [], next_invoice_number, messages

    except Exception as e:
        logger.error(f"Error processing {filepath}: {str(e)}")
        messages.append(('danger', f"خطا در پردازش فایل {os.path.basename(filepath)}: {str(e)}"))
        db.session.rollback()
        return pd.DataFrame(), [], next_invoice_number, messages

    return pd.DataFrame(output_data), log_entries, next_invoice_number, messages

def generate_sjt_output_excel(data_df, template_path, output_path):
    try:
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"فایل الگو در مسیر زیر یافت نشد: {template_path}")
        df = data_df.copy()

        logger.debug(f"Before multiplier - Q: {df['Q'].values if 'Q' in df else 'Not found'}, S: {df['S'].values if 'S' in df else 'Not found'}")

        if 'Q' in df.columns:
            df['Q'] = pd.to_numeric(df['Q'], errors='coerce') * MULTIPLIER_FACTOR
        if 'S' in df.columns:
            df['S'] = pd.to_numeric(df['S'], errors='coerce') * MULTIPLIER_FACTOR

        logger.debug(f"After multiplier - Q: {df['Q'].values if 'Q' in df else 'Not found'}, S: {df['S'].values if 'S' in df else 'Not found'}")

        book = load_workbook(template_path, keep_vba=True)
        sheet = book.active

        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))
        
        start_row = 2
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
            for cell in row:
                cell.value = None
        
        for index, row_data in df.iterrows():
            current_write_row = start_row + index
            for col_name, value in row_data.items():
                col_index = ord(col_name) - ord('A') + 1
                cell = sheet.cell(row=current_write_row, column=col_index, value=value)
        
        book.save(output_path)
        return output_path, None
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        return None, str(e)

def generate_usage_log_excel(data_df, output_path):
    try:
        data_df.to_excel(output_path, index=False, engine='openpyxl')
        return output_path, None
    except Exception as e:
        return None, str(e)

def calculate_inventory_values(db, Item, Settings):
    """
    Calculates initial, used, and remaining inventory values based on Item table.
    Updates these values in the Settings table.
    """
    try:
        results = db.session.query(
            func.sum(Item.unit_price * Item.quantity).label('initial_value'),
            func.sum(Item.unit_price * Item.remaining_quantity).label('remaining_value'),
            func.sum(Item.unit_price * (Item.quantity - Item.remaining_quantity)).label('used_value')
        ).one()

        initial_value = float(results.initial_value or 0)
        remaining_value = float(results.remaining_value or 0)
        used_value = float(results.used_value or 0)

        logger.debug(f"Calculated inventory values: Initial={initial_value}, Remaining={remaining_value}, Used={used_value}")

        settings = {
            'INITIAL_INVENTORY_VALUE': str(initial_value),
            'REMAINING_INVENTORY_VALUE': str(remaining_value),
            'USED_INVENTORY_VALUE': str(used_value)
        }

        for setting_name, setting_value in settings.items():
            setting = Settings.query.filter_by(setting_name=setting_name).first()
            if setting:
                setting.setting_value = setting_value
            else:
                setting = Settings(setting_name=setting_name, setting_value=setting_value)
                db.session.add(setting)

        db.session.commit()
        logger.debug("Inventory values updated in Settings table.")

        return initial_value, remaining_value, used_value

    except Exception as e:
        logger.error(f"Error calculating inventory values: {str(e)}")
        db.session.rollback()
        return 0, 0, 0